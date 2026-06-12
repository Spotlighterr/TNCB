import sys
import subprocess

# Ensure openpyxl is installed
try:
    import openpyxl
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("Installing openpyxl...")
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl"], check=True)
    import openpyxl
    from openpyxl.worksheet.datavalidation import DataValidation

def create_excel():
    wb = openpyxl.Workbook()
    
    # 1. Lists worksheet
    ws_lists = wb.active
    ws_lists.title = "Lists"
    
    # Hanoi districts in Column A
    hanoi_districts = [
        "Ba Đình", "Hoàn Kiếm", "Tây Hồ", "Long Biên", "Cầu Giấy", "Đống Đa", "Hai Bà Trưng",
        "Hoàng Mai", "Thanh Xuân", "Nam Từ Liêm", "Bắc Từ Liêm", "Hà Đông", "Gia Lâm", "Hoài Đức", "Thanh Trì"
    ]
    ws_lists.cell(row=1, column=1, value="Hà_Nội")
    for r_idx, d in enumerate(hanoi_districts, start=2):
        ws_lists.cell(row=r_idx, column=1, value=d)
        
    # HCMC districts in Column B
    hcmc_districts = [
        "Quận 1", "Quận 3", "Quận 4", "Quận 5", "Quận 6", "Quận 7", "Quận 8", "Quận 10", 
        "Quận 11", "Quận 12", "Bình Thạnh", "Gò Vấp", "Phú Nhuận", "Tân Bình", "Tân Phú", 
        "Bình Tân", "Thủ Đức", "Bình Chánh", "Hóc Môn", "Nhà Bè"
    ]
    ws_lists.cell(row=1, column=2, value="TP_Hồ_Chí_Minh")
    for r_idx, d in enumerate(hcmc_districts, start=2):
        ws_lists.cell(row=r_idx, column=2, value=d)
        
    # Cities in Column D
    cities = ["Hà Nội", "TP Hồ Chí Minh"]
    ws_lists.cell(row=1, column=4, value="Thành phố")
    for r_idx, c in enumerate(cities, start=2):
        ws_lists.cell(row=r_idx, column=4, value=c)
        
    # Room types in Column E
    room_types = ["Chung cư mini", "Căn hộ", "Phòng trọ", "Nhà nguyên căn", "Kí túc xá"]
    ws_lists.cell(row=1, column=5, value="Loại phòng")
    for r_idx, rt in enumerate(room_types, start=2):
        ws_lists.cell(row=r_idx, column=5, value=rt)

    # Register Named Ranges in Workbook
    # Hanoi Named Range (Lists!$A$2:$A$16)
    # openpyxl syntax: DefinedName(name, list of range expressions)
    from openpyxl.workbook.defined_name import DefinedName
    
    hn_range = f"Lists!$A$2:$A${len(hanoi_districts) + 1}"
    hn_def = DefinedName("Hà_Nội", attr_text=hn_range)
    wb.defined_names["Hà_Nội"] = hn_def
    
    hcm_range = f"Lists!$B$2:$B${len(hcmc_districts) + 1}"
    hcm_def = DefinedName("TP_Hồ_Chí_Minh", attr_text=hcm_range)
    wb.defined_names["TP_Hồ_Chí_Minh"] = hcm_def

    # 2. Properties worksheet
    ws_props = wb.create_sheet(title="Properties")
    
    headers = [
        "Tiêu đề", "Loại phòng", "Giá thuê", "Diện tích", "Thành phố", "Quận huyện", 
        "Phường xã", "Địa chỉ chi tiết", "Tọa độ", "Hình ảnh", "Đã Review", "Điện", 
        "Nước", "Dịch vụ", "Mô tả", "Điều hòa", "Nóng lạnh", "Wifi", "Máy giặt", 
        "Tủ lạnh", "Khóa vân tay", "Giờ giấc tự do", "Chỗ để xe", "Bếp riêng", "Bảo vệ 24/7"
    ]
    for col_idx, h in enumerate(headers, start=1):
        ws_props.cell(row=1, column=col_idx, value=h)
        
    # Add a sample row
    sample_row = [
        "Chung cư mini Cầu Giấy Full đồ ban công thoáng", "Chung cư mini", 4500000, 28, "Hà Nội", "Cầu Giấy", 
        "Dịch Vọng", "Số 12 Ngõ Chùa Hà", "", 
        "https://images.unsplash.com/photo-1522708323590-d24dbb6b0267?auto=format&fit=crop&w=800&q=80", 
        "yes", 3800, 100000, 150000, 
        "Chung cư mini khép kín full đồ đạc giường tủ đệm ga gối điều hòa nóng lạnh máy giặt tủ lạnh ban công phơi đồ rộng thoáng mát an ninh tốt khóa vân tay.",
        True, True, True, True, True, False, False, True, True, False
    ]
    for col_idx, val in enumerate(sample_row, start=1):
        ws_props.cell(row=2, column=col_idx, value=val)

    # 3. Add Data Validations
    # City Validation (Column E: "Thành phố")
    dv_city = DataValidation(type="list", formula1="Lists!$D$2:$D$3", allow_blank=True)
    ws_props.add_data_validation(dv_city)
    dv_city.add("E2:E200")
    
    # Room Type Validation (Column B: "Loại phòng")
    dv_type = DataValidation(type="list", formula1="Lists!$E$2:$E$6", allow_blank=True)
    ws_props.add_data_validation(dv_type)
    dv_type.add("B2:B200")
    
    # District Dependent Validation (Column F: "Quận huyện")
    # Formula uses E2 (city column of the row)
    dv_district = DataValidation(type="list", formula1='=INDIRECT(SUBSTITUTE(E2," ","_"))', allow_blank=True)
    ws_props.add_data_validation(dv_district)
    dv_district.add("F2:F200")

    wb.save("sample_properties.xlsx")
    print("sample_properties.xlsx created successfully.")

if __name__ == "__main__":
    create_excel()
